"""
Module 9 — Report Generation
Produces PDF and Excel reports for daily / weekly / monthly periods.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pandas as pd
from datetime import date, timedelta
from pathlib import Path

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import REPORTS_DIR
from database.models import get_attendance_df, get_security_logs_df, get_total_employees
from app.analytics import compute_kpis, generate_insights


ACCENT = colors.HexColor("#1a73e8")
LIGHT  = colors.HexColor("#e8f0fe")
GRAY   = colors.HexColor("#f5f5f5")


def _period_df(period: str):
    mapping = {"daily": 1, "weekly": 7, "monthly": 30}
    days = mapping.get(period, 1)
    df = get_attendance_df(days=days)
    return df, days


# ── PDF ───────────────────────────────────────────────────────

def generate_pdf_report(period: str = "daily") -> str:
    df, days = _period_df(period)
    kpis    = compute_kpis()
    insights = generate_insights(kpis)

    filename = REPORTS_DIR / f"attendance_{period}_{date.today().isoformat()}.pdf"
    doc = SimpleDocTemplate(str(filename), pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=ACCENT, fontSize=20, spaceAfter=6)
    story.append(Paragraph("Smart Attendance & Security Report", title_style))
    story.append(Paragraph(
        f"{period.capitalize()} Report  ·  Generated {date.today().strftime('%d %B %Y')}",
        styles["Normal"]
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=12))

    # KPI summary table
    story.append(Paragraph("Key Performance Indicators", styles["Heading2"]))
    kpi_data = [
        ["Metric", "Value"],
        ["Total Employees",        str(kpis["total_employees"])],
        ["Present Today",          str(kpis["present_today"])],
        ["Absent Today",           str(kpis["absent_today"])],
        ["Late Today",             str(kpis["late_today"])],
        ["Attendance Rate (30d)",  f"{kpis['attendance_rate_30d']}%"],
        ["Security Incidents (30d)", str(kpis["security_incidents_30d"])],
    ]
    kpi_table = Table(kpi_data, colWidths=[9*cm, 8*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), ACCENT),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 10),
        ("BACKGROUND", (0,1), (-1,-1), GRAY),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, GRAY]),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("ROWHEIGHT",  (0,0), (-1,-1), 20),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 12))

    # AI Insights
    story.append(Paragraph("AI-Generated Insights", styles["Heading2"]))
    for insight in insights:
        story.append(Paragraph(f"  {insight}", styles["Normal"]))
    story.append(Spacer(1, 12))

    # Attendance records
    if not df.empty:
        story.append(Paragraph("Attendance Records", styles["Heading2"]))
        display_df = df[["date", "full_name", "employee_id", "department",
                          "check_in_time", "status"]].head(100)
        headers = ["Date", "Name", "Employee ID", "Department", "Check-in", "Status"]
        data = [headers] + display_df.values.tolist()
        col_w = [2.5*cm, 4*cm, 2.5*cm, 3*cm, 2*cm, 2*cm]
        att_table = Table(data, colWidths=col_w, repeatRows=1)
        att_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), ACCENT),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, LIGHT]),
            ("GRID",       (0,0), (-1,-1), 0.3, colors.lightgrey),
            ("LEFTPADDING", (0,0), (-1,-1), 4),
            ("ROWHEIGHT",  (0,0), (-1,-1), 16),
        ]))
        story.append(att_table)

    doc.build(story)
    print(f"✅ PDF report saved: {filename}")
    return str(filename)


# ── Excel ─────────────────────────────────────────────────────

def generate_excel_report(period: str = "daily") -> str:
    df, days = _period_df(period)
    sec_df   = get_security_logs_df(days=days)
    kpis     = compute_kpis()
    insights = generate_insights(kpis)

    filename = REPORTS_DIR / f"attendance_{period}_{date.today().isoformat()}.xlsx"
    wb = openpyxl.Workbook()

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1A73E8")
    alt_fill     = PatternFill("solid", fgColor="E8F0FE")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Sheet 1: Summary ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Smart Attendance Report"
    ws1["A1"].font = Font(bold=True, size=16, color="1A73E8")
    ws1["A2"] = f"{period.capitalize()} · {date.today().strftime('%d %B %Y')}"
    ws1["A2"].font = Font(italic=True, size=11, color="666666")
    ws1.append([])

    ws1.append(["Metric", "Value"])
    style_header_row(ws1, ws1.max_row, 2)
    for key, val in [
        ("Total Employees",         kpis["total_employees"]),
        ("Present Today",           kpis["present_today"]),
        ("Absent Today",            kpis["absent_today"]),
        ("Late Today",              kpis["late_today"]),
        ("Attendance Rate (30d)",   f"{kpis['attendance_rate_30d']}%"),
        ("Security Incidents (30d)",kpis["security_incidents_30d"]),
    ]:
        ws1.append([key, val])

    ws1.append([])
    ws1.append(["AI Insights"])
    ws1.cell(ws1.max_row, 1).font = Font(bold=True, size=12, color="1A73E8")
    for insight in insights:
        ws1.append([insight])
    auto_width(ws1)

    # ── Sheet 2: Attendance ───────────────────────────────────
    if not df.empty:
        ws2 = wb.create_sheet("Attendance")
        cols = ["date", "full_name", "employee_id", "department", "check_in_time", "status"]
        headers = ["Date", "Name", "Employee ID", "Department", "Check-in Time", "Status"]
        ws2.append(headers)
        style_header_row(ws2, 1, len(headers))
        for i, (_, row) in enumerate(df[cols].iterrows(), start=2):
            ws2.append(list(row))
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws2.cell(i, col).fill = alt_fill
        auto_width(ws2)

    # ── Sheet 3: Security ─────────────────────────────────────
    if not sec_df.empty:
        ws3 = wb.create_sheet("Security Logs")
        s_cols = ["log_id", "image_path", "detected_at", "alert_sent", "notes"]
        s_headers = ["Log ID", "Image Path", "Detected At", "Alert Sent", "Notes"]
        ws3.append(s_headers)
        style_header_row(ws3, 1, len(s_headers))
        for i, (_, row) in enumerate(sec_df[s_cols].iterrows(), start=2):
            ws3.append(list(row))
        auto_width(ws3)

    wb.save(str(filename))
    print(f"✅ Excel report saved: {filename}")
    return str(filename)


if __name__ == "__main__":
    generate_pdf_report("monthly")
    generate_excel_report("monthly")
